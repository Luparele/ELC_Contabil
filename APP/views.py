from django.shortcuts import render, redirect, get_object_or_404
from .forms import DespesaForm, ReceitaForm, CategoriaForm, PerfilEmpresaForm, ContaBancariaForm, FornecedorForm, DASN_SIMEIForm
from .models import Despesa, Receita, Categoria, PerfilEmpresa, ContaBancaria, DeclaracaoAnual, Fornecedor, PreferenciaUsuario, DASN_SIMEI
from django.contrib.auth.models import User
from django.contrib import messages
from itertools import chain
from operator import attrgetter
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
import datetime
from django.db.models import Sum, Q
import json
import calendar
import requests
from django.http import HttpResponse, JsonResponse
import csv
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

# --- NOVAS IMPORTAÇÕES CORRIGIDAS ---
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4 # Corrigido para 'A4' maiúsculo
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors

@login_required
def dashboard(request):
    hoje = datetime.date.today()
    ano_corrente = hoje.year
    ano_anterior = ano_corrente - 1
    
    receitas_mes_atual = Receita.objects.filter(usuario=request.user, data__year=hoje.year, data__month=hoje.month)
    despesas_mes_atual = Despesa.objects.filter(usuario=request.user, data__year=hoje.year, data__month=hoje.month)
    total_receitas = receitas_mes_atual.aggregate(Sum('valor'))['valor__sum'] or 0
    total_despesas = despesas_mes_atual.aggregate(Sum('valor'))['valor__sum'] or 0
    balanco = total_receitas - total_despesas

    faturamento_anual = Receita.objects.filter(usuario=request.user, data__year=ano_corrente).aggregate(Sum('valor'))['valor__sum'] or 0
    faturamento_ano_anterior = Receita.objects.filter(usuario=request.user, data__year=ano_anterior).aggregate(Sum('valor'))['valor__sum'] or 0

    alerta_ano_anterior = False
    if hoje.year > ano_anterior and faturamento_ano_anterior > 0:
        # CORREÇÃO: Tentar obter o perfil, mas não falhar se não existir
        try:
            perfil = request.user.perfilempresa
            declaracao_feita = DeclaracaoAnual.objects.filter(perfil_empresa=perfil, ano=ano_anterior).exists()
            if not declaracao_feita:
                alerta_ano_anterior = True
        except PerfilEmpresa.DoesNotExist:
            alerta_ano_anterior = False # Não pode ter alerta se o perfil não existe
    
    # === SISTEMA DE ALERTAS ===
    alertas = []
    preferencias, _ = PreferenciaUsuario.objects.get_or_create(usuario=request.user)
    
    if preferencias.alertas_ativos:
        # Alerta 1: Despesas acima do percentual definido
        if total_receitas > 0:
            percentual_despesas = (total_despesas / total_receitas) * 100
            if percentual_despesas >= preferencias.alerta_percentual_despesas:
                alertas.append({
                    'tipo': 'warning',
                    'icone': 'bi-exclamation-triangle-fill',
                    'titulo': 'Despesas Elevadas',
                    'mensagem': f'Suas despesas estão em {percentual_despesas:.1f}% das receitas do mês (limite: {preferencias.alerta_percentual_despesas}%).'
                })
        
        # Alerta 2: Balanço negativo no mês
        if balanco < 0:
            alertas.append({
                'tipo': 'danger',
                'icone': 'bi-x-circle-fill',
                'titulo': 'Balanço Negativo',
                'mensagem': f'O balanço do mês está negativo em R$ {abs(balanco):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            })
        
        # Alerta 3: Comparação com mês anterior
        mes_anterior = hoje.month - 1 if hoje.month > 1 else 12
        ano_mes_anterior = hoje.year if hoje.month > 1 else hoje.year - 1
        
        despesas_mes_anterior = Despesa.objects.filter(
            usuario=request.user, 
            data__year=ano_mes_anterior, 
            data__month=mes_anterior
        ).aggregate(Sum('valor'))['valor__sum'] or 0
        
        if despesas_mes_anterior > 0 and total_despesas > 0:
            variacao = ((total_despesas - despesas_mes_anterior) / despesas_mes_anterior) * 100
            if variacao > 20:  # Aumento de mais de 20%
                alertas.append({
                    'tipo': 'info',
                    'icone': 'bi-graph-up-arrow',
                    'titulo': 'Aumento nas Despesas',
                    'mensagem': f'Suas despesas aumentaram {variacao:.1f}% em relação ao mês passado.'
                })
        
        # Alerta 4: Lançamentos sem categoria (últimos 30 dias)
        data_30_dias = hoje - datetime.timedelta(days=30)
        lancamentos_sem_categoria = (
            Receita.objects.filter(usuario=request.user, data__gte=data_30_dias, categoria__isnull=True).count() +
            Despesa.objects.filter(usuario=request.user, data__gte=data_30_dias, categoria__isnull=True).count()
        )
        
        if lancamentos_sem_categoria > 0:
            alertas.append({
                'tipo': 'warning',
                'icone': 'bi-tag',
                'titulo': 'Lançamentos sem Categoria',
                'mensagem': f'Você tem {lancamentos_sem_categoria} lançamento(s) sem categoria nos últimos 30 dias.'
            })
        
        # Alerta 5: Fornecedores inativos com lançamentos recentes
        fornecedores_inativos_com_lancamentos = Fornecedor.objects.filter(
            usuario=request.user,
            ativo=False
        ).filter(
            Q(receitas__data__gte=data_30_dias) | Q(despesas__data__gte=data_30_dias)
        ).distinct().count()
        
        if fornecedores_inativos_com_lancamentos > 0:
            alertas.append({
                'tipo': 'info',
                'icone': 'bi-person-x',
                'titulo': 'Fornecedores Inativos',
                'mensagem': f'{fornecedores_inativos_com_lancamentos} fornecedor(es) inativo(s) possui(em) lançamentos recentes.'
            })
    # === FIM SISTEMA DE ALERTAS ===

    labels_grafico, dados_receitas, dados_despesas = [], [], []
    for i in range(6):
        mes, ano = (hoje.month - i, hoje.year)
        if mes <= 0: mes += 12; ano -= 1
        nome_mes_pt = calendar.month_name[mes][:3].capitalize()
        labels_grafico.append(f'{nome_mes_pt}/{ano}')
        receitas_do_mes = Receita.objects.filter(usuario=request.user, data__year=ano, data__month=mes).aggregate(Sum('valor'))['valor__sum'] or 0
        dados_receitas.append(float(receitas_do_mes))
        despesas_do_mes = Despesa.objects.filter(usuario=request.user, data__year=ano, data__month=mes).aggregate(Sum('valor'))['valor__sum'] or 0
        dados_despesas.append(float(despesas_do_mes))
    labels_grafico.reverse(); dados_receitas.reverse(); dados_despesas.reverse()

    # --- INÍCIO: NOVOS CÁLCULOS PARA GRÁFICOS PIZZA ---

    # 1. Gráfico Pizza Mensal
    despesas_mes_categoria = despesas_mes_atual.values('categoria__nome') \
        .annotate(total=Sum('valor')) \
        .order_by('-total')
    
    labels_pie_mes = []
    dados_pie_mes = []
    for item in despesas_mes_categoria:
        labels_pie_mes.append(item['categoria__nome'] or 'Sem Categoria')
        dados_pie_mes.append(float(item['total']))

    # 2. Gráfico Pizza Anual
    despesas_ano_categoria = Despesa.objects.filter(usuario=request.user, data__year=hoje.year) \
        .values('categoria__nome') \
        .annotate(total=Sum('valor')) \
        .order_by('-total')
    
    labels_pie_ano = []
    dados_pie_ano = []
    for item in despesas_ano_categoria:
        labels_pie_ano.append(item['categoria__nome'] or 'Sem Categoria')
        dados_pie_ano.append(float(item['total']))

    # --- FIM: NOVOS CÁLCULOS ---

    # === NOVOS GRÁFICOS ADICIONAIS ===
    
    # 3. Top 5 Maiores Despesas do Mês
    top5_despesas = despesas_mes_atual.order_by('-valor')[:5]
    labels_top5 = [d.descricao[:30] for d in top5_despesas]
    valores_top5 = [float(d.valor) for d in top5_despesas]
    
    # 4. Comparação Mês Atual vs Anterior
    mes_anterior = hoje.month - 1 if hoje.month > 1 else 12
    ano_mes_anterior = hoje.year if hoje.month > 1 else hoje.year - 1
    
    receitas_mes_anterior = Receita.objects.filter(
        usuario=request.user, 
        data__year=ano_mes_anterior, 
        data__month=mes_anterior
    ).aggregate(Sum('valor'))['valor__sum'] or 0
    
    despesas_mes_anterior = Despesa.objects.filter(
        usuario=request.user, 
        data__year=ano_mes_anterior, 
        data__month=mes_anterior
    ).aggregate(Sum('valor'))['valor__sum'] or 0
    
    balanco_mes_anterior = receitas_mes_anterior - despesas_mes_anterior
    
    # 5. Evolução do Balanço (6 meses)
    labels_balanco = []
    dados_balanco = []
    for i in range(6):
        mes, ano = (hoje.month - i, hoje.year)
        if mes <= 0: 
            mes += 12
            ano -= 1
        nome_mes_pt = calendar.month_name[mes][:3].capitalize()
        labels_balanco.append(f'{nome_mes_pt}/{ano}')
        
        rec_mes = Receita.objects.filter(
            usuario=request.user, 
            data__year=ano, 
            data__month=mes
        ).aggregate(Sum('valor'))['valor__sum'] or 0
        
        desp_mes = Despesa.objects.filter(
            usuario=request.user, 
            data__year=ano, 
            data__month=mes
        ).aggregate(Sum('valor'))['valor__sum'] or 0
        
        dados_balanco.append(float(rec_mes - desp_mes))
    
    labels_balanco.reverse()
    dados_balanco.reverse()
    
    # === FIM NOVOS GRÁFICOS ===

    admin_data = {}
    if request.user.is_staff:
        TAMANHO_TOTAL_DISCO_MB, TAMANHO_SISTEMA_MB = 5120, 292.1
        
        consumo_por_usuario, total_uploads_bytes = [], 0
        todos_usuarios = User.objects.all().order_by('username')

        for user in todos_usuarios:
            consumo_usuario_bytes = 0
            for receita in user.receita_set.all():
                if receita.comprovante:
                    try: consumo_usuario_bytes += receita.comprovante.size
                    except FileNotFoundError: pass
            for despesa in user.despesa_set.all():
                if despesa.comprovante:
                    try: consumo_usuario_bytes += despesa.comprovante.size
                    except FileNotFoundError: pass
            consumo_por_usuario.append({'username': user.username, 'is_staff': user.is_staff, 'consumo_bytes': consumo_usuario_bytes})
            total_uploads_bytes += consumo_usuario_bytes

        total_uploads_mb = total_uploads_bytes / (1024 * 1024)
        
        percentual_sistema = (TAMANHO_SISTEMA_MB / TAMANHO_TOTAL_DISCO_MB) * 100 if TAMANHO_TOTAL_DISCO_MB > 0 else 0
        percentual_uploads = (total_uploads_mb / TAMANHO_TOTAL_DISCO_MB) * 100 if TAMANHO_TOTAL_DISCO_MB > 0 else 0
        
        for item in consumo_por_usuario:
            consumo_mb = item['consumo_bytes'] / (1024 * 1024)
            item['consumo_mb'] = consumo_mb
            item['percentual_do_total'] = (consumo_mb / TAMANHO_TOTAL_DISCO_MB) * 100 if TAMANHO_TOTAL_DISCO_MB > 0 else 0
        
        consumo_por_usuario.sort(key=lambda x: x['consumo_bytes'], reverse=True)

        admin_data = {
            'TAMANHO_TOTAL_DISCO_MB': TAMANHO_TOTAL_DISCO_MB,
            'TAMANHO_TOTAL_DISCO_DISPLAY': '5 GB',
            'total_geral_mb': TAMANHO_SISTEMA_MB + total_uploads_mb,
            'consumo_por_usuario': consumo_por_usuario,
            'percentual_sistema_css': f"{percentual_sistema:.2f}".replace(',', '.'),
            'percentual_uploads_css': f"{percentual_uploads:.2f}".replace(',', '.'),
            'TAMANHO_SISTEMA_MB': TAMANHO_SISTEMA_MB,
            'total_uploads_mb': total_uploads_mb,
        }

    context = {
        'total_receitas': total_receitas, 'total_despesas': total_despesas, 'balanco': balanco,
        'faturamento_anual': faturamento_anual, 'faturamento_ano_anterior': faturamento_ano_anterior,
        'ano_corrente': ano_corrente, 'ano_anterior': ano_anterior, 'alerta_ano_anterior': alerta_ano_anterior,
        'labels_grafico': json.dumps(labels_grafico), 'dados_receitas': json.dumps(dados_receitas),
        'dados_despesas': json.dumps(dados_despesas), 'admin_data': admin_data,

        # --- NOVAS VARIÁVEIS ADICIONADAS AO CONTEXTO ---
        'labels_pie_mes_json': json.dumps(labels_pie_mes),
        'dados_pie_mes_json': json.dumps(dados_pie_mes),
        'labels_pie_ano_json': json.dumps(labels_pie_ano),
        'dados_pie_ano_json': json.dumps(dados_pie_ano),
        
        # ALERTAS
        'alertas': alertas,
        
        # TEMA
        'tema_escuro': preferencias.tema_escuro,
        
        # NOVOS GRÁFICOS
        'labels_top5_json': json.dumps(labels_top5),
        'valores_top5_json': json.dumps(valores_top5),
        'receitas_mes_anterior': receitas_mes_anterior,
        'despesas_mes_anterior': despesas_mes_anterior,
        'balanco_mes_anterior': balanco_mes_anterior,
        'labels_balanco_json': json.dumps(labels_balanco),
        'dados_balanco_json': json.dumps(dados_balanco),
    }
    return render(request, 'APP/dashboard.html', context)

@login_required
def confirmar_declaracao(request, ano):
    perfil = request.user.perfilempresa
    DeclaracaoAnual.objects.get_or_create(perfil_empresa=perfil, ano=ano)
    messages.success(request, f'Confirmação da declaração do ano {ano} registrada com sucesso!')
    return redirect('dashboard')

@login_required
def toggle_tema(request):
    """Alterna o tema escuro e salva a preferência"""
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        tema_escuro = data.get('tema_escuro', False)
        
        preferencias, _ = PreferenciaUsuario.objects.get_or_create(usuario=request.user)
        preferencias.tema_escuro = tema_escuro
        preferencias.save()
        
        return JsonResponse({'status': 'success', 'tema_escuro': tema_escuro})
    
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def listar_lancamentos(request):
    """Lista lançamentos com paginação e filtros avançados"""
    
    # Obter preferências do usuário
    preferencias, _ = PreferenciaUsuario.objects.get_or_create(usuario=request.user)
    itens_por_pagina = int(request.GET.get('per_page', preferencias.itens_por_pagina))
    
    # Buscar lançamentos
    receitas = Receita.objects.filter(usuario=request.user).select_related('categoria', 'fornecedor')
    despesas = Despesa.objects.filter(usuario=request.user).select_related('categoria', 'fornecedor')
    
    # Filtros
    busca = request.GET.get('busca', '')
    tipo = request.GET.get('tipo', '')  # R, D ou vazio
    categoria_id = request.GET.get('categoria', '')
    fornecedor_id = request.GET.get('fornecedor', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    valor_min = request.GET.get('valor_min', '')
    valor_max = request.GET.get('valor_max', '')
    
    # Aplicar filtros
    if busca:
        receitas = receitas.filter(Q(descricao__icontains=busca) | Q(observacoes__icontains=busca))
        despesas = despesas.filter(Q(descricao__icontains=busca) | Q(observacoes__icontains=busca))
    
    if categoria_id:
        receitas = receitas.filter(categoria_id=categoria_id)
        despesas = despesas.filter(categoria_id=categoria_id)
    
    if fornecedor_id:
        receitas = receitas.filter(fornecedor_id=fornecedor_id)
        despesas = despesas.filter(fornecedor_id=fornecedor_id)
    
    if data_inicio:
        receitas = receitas.filter(data__gte=data_inicio)
        despesas = despesas.filter(data__gte=data_inicio)
    
    if data_fim:
        receitas = receitas.filter(data__lte=data_fim)
        despesas = despesas.filter(data__lte=data_fim)
    
    if valor_min:
        receitas = receitas.filter(valor__gte=valor_min)
        despesas = despesas.filter(valor__gte=valor_min)
    
    if valor_max:
        receitas = receitas.filter(valor__lte=valor_max)
        despesas = despesas.filter(valor__lte=valor_max)
    
    # Filtrar por tipo
    if tipo == 'R':
        todos_lancamentos = list(receitas)
    elif tipo == 'D':
        todos_lancamentos = list(despesas)
    else:
        todos_lancamentos = sorted(
            chain(receitas, despesas),
            key=attrgetter('data'),
            reverse=True
        )
    
    # Paginação
    paginator = Paginator(todos_lancamentos, itens_por_pagina)
    page = request.GET.get('page', 1)
    
    try:
        lancamentos_paginados = paginator.page(page)
    except PageNotAnInteger:
        lancamentos_paginados = paginator.page(1)
    except EmptyPage:
        lancamentos_paginados = paginator.page(paginator.num_pages)
    
    # Dados para filtros
    categorias = Categoria.objects.filter(
        Q(usuario=request.user) | Q(usuario__isnull=True, is_padrao=True)
    ).filter(ativo=True).order_by('tipo', 'nome')
    
    fornecedores = Fornecedor.objects.filter(
        usuario=request.user, 
        ativo=True
    ).order_by('nome')
    
    context = {
        'lancamentos': lancamentos_paginados,
        'categorias': categorias,
        'fornecedores': fornecedores,
        'filtros': {
            'busca': busca,
            'tipo': tipo,
            'categoria_id': categoria_id,
            'fornecedor_id': fornecedor_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'valor_min': valor_min,
            'valor_max': valor_max,
            'per_page': itens_por_pagina,
        },
        'total_lancamentos': len(todos_lancamentos),
    }
    return render(request, 'APP/lancamento_list.html', context)

@login_required
def adicionar_despesa(request):
    if request.method == 'POST':
        form = DespesaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.usuario = request.user
            
            # Criar fornecedor se preencheu o campo novo_fornecedor
            novo_fornecedor_nome = form.cleaned_data.get('novo_fornecedor')
            if novo_fornecedor_nome:
                fornecedor = Fornecedor.objects.create(
                    usuario=request.user,
                    nome=novo_fornecedor_nome,
                    tipo='PJ'
                )
                despesa.fornecedor = fornecedor
            
            despesa.save()
            messages.success(request, 'Despesa salva com sucesso!')
            return redirect('listar_lancamentos')
    else:
        form = DespesaForm(user=request.user)
    return render(request, 'APP/despesa_form.html', {'form': form})

@login_required
def adicionar_receita(request):
    if request.method == 'POST':
        form = ReceitaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            receita = form.save(commit=False)
            receita.usuario = request.user
            
            # Criar cliente se preencheu o campo novo_fornecedor
            novo_fornecedor_nome = form.cleaned_data.get('novo_fornecedor')
            if novo_fornecedor_nome:
                fornecedor = Fornecedor.objects.create(
                    usuario=request.user,
                    nome=novo_fornecedor_nome,
                    tipo='PF'  # Pessoa Física para clientes
                )
                receita.fornecedor = fornecedor
            
            receita.save()
            messages.success(request, 'Receita salva com sucesso!')
            return redirect('listar_lancamentos')
    else:
        form = ReceitaForm(user=request.user)
    return render(request, 'APP/receita_form.html', {'form': form})

@login_required
def relatorios(request):
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')
    tipo_lancamento = request.GET.get('tipo_lancamento')
    categoria_id = request.GET.get('categoria')

    filtros_aplicados = 'action' in request.GET

    lancamentos_filtrados = []
    total_receitas_periodo = 0
    total_despesas_periodo = 0
    balanco_periodo = 0

    if filtros_aplicados:
        data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else None
        data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else None

        receitas = Receita.objects.filter(usuario=request.user)
        despesas = Despesa.objects.filter(usuario=request.user)
        if data_inicio:
            receitas = receitas.filter(data__gte=data_inicio)
            despesas = despesas.filter(data__gte=data_inicio)
        if data_fim:
            receitas = receitas.filter(data__lte=data_fim)
            despesas = despesas.filter(data__lte=data_fim)
        if categoria_id:
            receitas = receitas.filter(categoria_id=categoria_id)
            despesas = despesas.filter(categoria_id=categoria_id)

        lista_final = []
        if tipo_lancamento == 'R':
            lista_final = list(receitas)
        elif tipo_lancamento == 'D':
            lista_final = list(despesas)
        else:
            lista_final = list(chain(receitas, despesas))

        lista_final.sort(key=attrgetter('data'), reverse=True)
        lancamentos_filtrados = lista_final

        total_receitas_periodo = sum(l.valor for l in lancamentos_filtrados if isinstance(l, Receita))
        total_despesas_periodo = sum(l.valor for l in lancamentos_filtrados if isinstance(l, Despesa))
        balanco_periodo = total_receitas_periodo - total_despesas_periodo

    categorias_por_tipo = {
        'R': list(Categoria.objects.filter(tipo='R').values('id', 'nome')),
        'D': list(Categoria.objects.filter(tipo='D').values('id', 'nome'))
    }

    context = {
        'lancamentos': lancamentos_filtrados,
        'total_receitas': total_receitas_periodo,
        'total_despesas': total_despesas_periodo,
        'balanco': balanco_periodo,
        'categorias_json': json.dumps(categorias_por_tipo),
        'filtros_aplicados': filtros_aplicados,
        'data_inicio_value': data_inicio_str,
        'data_fim_value': data_fim_str,
        'tipo_lancamento_value': tipo_lancamento,
        'categoria_id_value': categoria_id,
        # Variáveis de data formatada para o cabeçalho
        'data_inicio_fmt': datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').strftime('%d/%m/%Y') if data_inicio_str else "Início",
        'data_fim_fmt': datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').strftime('%d/%m/%Y') if data_fim_str else "Fim",
    }
    return render(request, 'APP/relatorios.html', context)

@login_required
def exportar_excel(request):
    """Exporta lançamentos para Excel formatado"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    # Obter filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_lancamento = request.GET.get('tipo_lancamento')
    categoria_id = request.GET.get('categoria')
    
    # Query base
    receitas_query = Receita.objects.filter(usuario=request.user)
    despesas_query = Despesa.objects.filter(usuario=request.user)
    
    # Aplicar filtros
    if data_inicio:
        receitas_query = receitas_query.filter(data__gte=data_inicio)
        despesas_query = despesas_query.filter(data__gte=data_inicio)
    if data_fim:
        receitas_query = receitas_query.filter(data__lte=data_fim)
        despesas_query = despesas_query.filter(data__lte=data_fim)
    if categoria_id:
        receitas_query = receitas_query.filter(categoria_id=categoria_id)
        despesas_query = despesas_query.filter(categoria_id=categoria_id)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho
    headers = ['Data', 'Tipo', 'Descrição', 'Categoria', 'Fornecedor', 'CPF/CNPJ', 'Valor']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    row = 2
    total_receitas = 0
    total_despesas = 0
    
    # Adicionar receitas
    if tipo_lancamento != 'D':
        for receita in receitas_query.select_related('categoria', 'fornecedor').order_by('data'):
            ws.cell(row=row, column=1, value=receita.data.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value='Receita')
            ws.cell(row=row, column=3, value=receita.descricao)
            ws.cell(row=row, column=4, value=receita.categoria.nome if receita.categoria else '-')
            ws.cell(row=row, column=5, value=receita.fornecedor.nome if receita.fornecedor else '-')
            ws.cell(row=row, column=6, value=receita.fornecedor.cpf_cnpj if receita.fornecedor else '-')
            ws.cell(row=row, column=7, value=float(receita.valor))
            
            total_receitas += float(receita.valor)
            
            # Estilo receita (verde)
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 7:
                    cell.number_format = 'R$ #,##0.00'
                    cell.font = Font(color="198754", bold=True)
            row += 1
    
    # Adicionar despesas
    if tipo_lancamento != 'R':
        for despesa in despesas_query.select_related('categoria', 'fornecedor').order_by('data'):
            ws.cell(row=row, column=1, value=despesa.data.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value='Despesa')
            ws.cell(row=row, column=3, value=despesa.descricao)
            ws.cell(row=row, column=4, value=despesa.categoria.nome if despesa.categoria else '-')
            ws.cell(row=row, column=5, value=despesa.fornecedor.nome if despesa.fornecedor else '-')
            ws.cell(row=row, column=6, value=despesa.fornecedor.cpf_cnpj if despesa.fornecedor else '-')
            ws.cell(row=row, column=7, value=-float(despesa.valor))
            
            total_despesas += float(despesa.valor)
            
            # Estilo despesa (vermelho)
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 7:
                    cell.number_format = 'R$ #,##0.00'
                    cell.font = Font(color="DC3545", bold=True)
            row += 1
    
    # Resumo financeiro
    row += 1
    ws.cell(row=row, column=6, value='Total Receitas:').font = Font(bold=True, color="198754")
    ws.cell(row=row, column=7, value=total_receitas).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, color="198754")
    
    row += 1
    ws.cell(row=row, column=6, value='Total Despesas:').font = Font(bold=True, color="DC3545")
    ws.cell(row=row, column=7, value=-total_despesas).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, color="DC3545")
    
    row += 1
    balanco = total_receitas - total_despesas
    ws.cell(row=row, column=6, value='BALANÇO:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=7, value=balanco).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=7).font = Font(bold=True, size=12, color="198754" if balanco >= 0 else "DC3545")
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    
    # Resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'lancamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@login_required
def exportar_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="relatorio_contabil_{datetime.date.today()}.csv"'
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_lancamento = request.GET.get('tipo_lancamento')
    categoria_id = request.GET.get('categoria')

    receitas = Receita.objects.filter(usuario=request.user)
    despesas = Despesa.objects.filter(usuario=request.user)

    if data_inicio:
        receitas = receitas.filter(data__gte=data_inicio)
        despesas = despesas.filter(data__gte=data_inicio)
    if data_fim:
        receitas = receitas.filter(data__lte=data_fim)
        despesas = despesas.filter(data__lte=data_fim)
    if categoria_id:
        receitas = receitas.filter(categoria_id=categoria_id)
        despesas = despesas.filter(categoria_id=categoria_id)

    if tipo_lancamento == 'R':
        lancamentos_filtrados = list(receitas)
    elif tipo_lancamento == 'D':
        lancamentos_filtrados = list(despesas)
    else:
        lancamentos_filtrados = list(chain(receitas, despesas))

    lancamentos_filtrados.sort(key=attrgetter('data'), reverse=True)

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Data', 'Descricao', 'Tipo', 'Categoria', 'Fornecedor', 'CNPJ/CPF', 'Valor'])

    for lancamento in lancamentos_filtrados:
        tipo = 'Receita' if isinstance(lancamento, Receita) else 'Despesa'
        valor_formatado = str(lancamento.valor).replace('.', ',')
        writer.writerow([
            lancamento.data.strftime('%d/%m/%Y'),
            lancamento.descricao,
            tipo,
            lancamento.categoria.nome if lancamento.categoria else 'Sem Categoria',
            lancamento.fornecedor.nome if lancamento.fornecedor else '-',
            lancamento.fornecedor.cpf_cnpj if lancamento.fornecedor else '-',
            valor_formatado
        ])
    return response

@login_required
def listar_categorias(request):
    categorias = Categoria.objects.all().order_by('nome')
    context = {
        'categorias': categorias
    }
    return render(request, 'APP/categoria_list.html', context)

@login_required
def adicionar_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria adicionada com sucesso!')
            return redirect('listar_categorias')
    else:
        form = CategoriaForm()
    context = {
        'form': form,
        'titulo': 'Adicionar Nova Categoria'
    }
    return render(request, 'APP/categoria_form.html', context)

@login_required
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoria atualizada com sucesso!')
            return redirect('listar_categorias')
    else:
        form = CategoriaForm(instance=categoria)
    context = {
        'form': form,
        'titulo': 'Editar Categoria'
    }
    return render(request, 'APP/categoria_form.html', context)

@login_required
def excluir_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoria excluída com sucesso!')
        return redirect('listar_categorias')
    context = {
        'categoria': categoria
    }
    return render(request, 'APP/categoria_confirm_delete.html', context)

@login_required
def ver_perfil(request):
    # Trocado .get() por .get_or_create() para evitar o erro 500
    perfil, created = PerfilEmpresa.objects.get_or_create(usuario=request.user)
    context = {
        'perfil': perfil
    }
    return render(request, 'APP/perfil_ver.html', context)

@login_required
def editar_perfil(request):
    perfil, created = PerfilEmpresa.objects.get_or_create(usuario=request.user)
    if request.method == 'POST':
        form = PerfilEmpresaForm(request.POST, instance=perfil)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil da empresa atualizado com sucesso!')
            return redirect('ver_perfil')
    else:
        form = PerfilEmpresaForm(instance=perfil)
    context = {
        'form': form
    }
    return render(request, 'APP/perfil_editar.html', context)

@login_required
def consultar_cnpj(request):
    cnpj = request.GET.get('cnpj', '')
    cnpj = ''.join(filter(str.isdigit, cnpj))
    if not cnpj:
        return JsonResponse({'error': 'CNPJ não fornecido'}, status=400)
    try:
        response = requests.get(f'https://brasilapi.com.br/api/cnpj/v1/{cnpj}')
        response.raise_for_status()
        data = response.json()
        return JsonResponse(data)
    except requests.RequestException as e:
        return JsonResponse({'error': f'Falha ao consultar o CNPJ: {e}'}, status=500)
    except ValueError:
        return JsonResponse({'error': 'Resposta inválida da API'}, status=500)

@login_required
def adicionar_conta(request):
    perfil = request.user.perfilempresa
    if request.method == 'POST':
        form = ContaBancariaForm(request.POST)
        if form.is_valid():
            conta = form.save(commit=False)
            conta.perfil_empresa = perfil
            conta.save()
            messages.success(request, 'Nova conta bancária adicionada com sucesso!')
            return redirect('ver_perfil')
    else:
        form = ContaBancariaForm()
    context = {
        'form': form,
        'titulo': 'Adicionar Nova Conta Bancária'
    }
    return render(request, 'APP/conta_form.html', context)

@login_required
def editar_conta(request, pk):
    conta = get_object_or_404(ContaBancaria, pk=pk, perfil_empresa=request.user.perfilempresa)
    if request.method == 'POST':
        form = ContaBancariaForm(request.POST, instance=conta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conta bancária atualizada com sucesso!')
            return redirect('ver_perfil')
    else:
        form = ContaBancariaForm(instance=conta)
    context = {
        'form': form,
        'titulo': 'Editar Conta Bancária'
    }
    return render(request, 'APP/conta_form.html', context)

@login_required
def excluir_conta(request, pk):
    conta = get_object_or_404(ContaBancaria, pk=pk, perfil_empresa=request.user.perfilempresa)
    if request.method == 'POST':
        conta.delete()
        messages.success(request, 'Conta bancária excluída com sucesso!')
        return redirect('ver_perfil')
    context = {
        'conta': conta
    }
    return render(request, 'APP/conta_confirm_delete.html', context)


@login_required
def exportar_pdf(request):
    """Gera relatório PDF em paisagem com paginação correta"""
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    
    # Coleta de dados
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')
    tipo_lancamento = request.GET.get('tipo_lancamento')
    categoria_id = request.GET.get('categoria')

    data_inicio = datetime.datetime.strptime(data_inicio_str, '%Y-%m-%d').date() if data_inicio_str else None
    data_fim = datetime.datetime.strptime(data_fim_str, '%Y-%m-%d').date() if data_fim_str else None

    receitas = Receita.objects.filter(usuario=request.user)
    despesas = Despesa.objects.filter(usuario=request.user)

    if data_inicio:
        receitas = receitas.filter(data__gte=data_inicio)
        despesas = despesas.filter(data__gte=data_inicio)
    if data_fim:
        receitas = receitas.filter(data__lte=data_fim)
        despesas = despesas.filter(data__lte=data_fim)
    if categoria_id:
        receitas = receitas.filter(categoria_id=categoria_id)
        despesas = despesas.filter(categoria_id=categoria_id)

    if tipo_lancamento == 'R':
        lista_final = list(receitas)
    elif tipo_lancamento == 'D':
        lista_final = list(despesas)
    else:
        lista_final = list(chain(receitas, despesas))
    
    lista_final.sort(key=attrgetter('data'), reverse=True)
    
    perfil = request.user.perfilempresa
    contas = perfil.contas.all().order_by('-preferencial', 'nome_banco')
    
    total_receitas = sum(l.valor for l in lista_final if isinstance(l, Receita))
    total_despesas = sum(l.valor for l in lista_final if isinstance(l, Despesa))
    balanco = total_receitas - total_despesas

    # Configuração PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"relatorio_{perfil.razao_social or 'empresa'}_{datetime.date.today()}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pagesize = landscape(A4)
    p = canvas.Canvas(response, pagesize=pagesize)
    width, height = pagesize
    
    cor_verde = colors.Color(0, 0.6, 0)
    cor_vermelha = colors.Color(0.8, 0, 0)
    cor_cinza = colors.Color(0.3, 0.3, 0.3)
    
    MARGEM = 1.5 * cm
    
    def cabecalho(y, pag):
        p.setFont("Helvetica-Bold", 10)
        p.drawString(MARGEM, y, "RELATÓRIO FINANCEIRO")
        p.drawRightString(width - MARGEM, y, f"Pág. {pag}")
        y -= 0.4 * cm
        
        p.setFont("Helvetica", 7)
        info = f"{perfil.razao_social or 'Empresa'}"
        if perfil.cnpj:
            info += f" - CNPJ: {perfil.cnpj}"
        p.drawString(MARGEM, y, info)
        
        if data_inicio and data_fim:
            periodo = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        elif data_inicio:
            periodo = f"Desde {data_inicio.strftime('%d/%m/%Y')}"
        elif data_fim:
            periodo = f"Até {data_fim.strftime('%d/%m/%Y')}"
        else:
            periodo = "Todos os registros"
        p.drawRightString(width - MARGEM, y, periodo)
        y -= 0.3 * cm
        
        p.setStrokeColor(colors.grey)
        p.line(MARGEM, y, width - MARGEM, y)
        return y - 0.3 * cm
    
    def rodape():
        p.setFont("Helvetica", 6)
        p.drawCentredString(width / 2, 0.7 * cm, 
            f"Emitido: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')} | ELC Contábil")
    
    # Primeira página
    pagina = 1
    y = height - 1 * cm
    y = cabecalho(y, pagina)
    
    # Resumo
    p.setFont("Helvetica", 7)
    p.setFillColor(cor_verde)
    p.drawString(MARGEM, y, f"Receitas: R$ {total_receitas:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    p.setFillColor(cor_vermelha)
    p.drawString(MARGEM + 5*cm, y, f"Despesas: R$ {total_despesas:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    p.setFillColor(cor_verde if balanco >= 0 else cor_vermelha)
    p.drawString(MARGEM + 10*cm, y, f"Balanço: R$ {balanco:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    p.setFillColor(colors.black)
    y -= 0.6 * cm
    
    # Contas (máx 2)
    if contas.exists():
        p.setFont("Helvetica", 6)
        for conta in contas[:2]:
            p.drawString(MARGEM, y, f"• {conta.nome_banco} Ag:{conta.agencia} CC:{conta.conta_corrente}")
            y -= 0.25 * cm
        y -= 0.2 * cm
    
    p.setFont("Helvetica-Bold", 7)
    p.drawString(MARGEM, y, f"Lançamentos: {len(lista_final)}")
    y -= 0.35 * cm
    
    # Preparar tabela
    dados = [['Data', 'Categoria', 'Fornecedor', 'CNPJ/CPF', 'Valor']]
    for lanc in lista_final:
        valor_fmt = f"{lanc.valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        prefixo = "+ " if isinstance(lanc, Receita) else "- "
        
        cat = (lanc.categoria.nome[:23]+'..') if lanc.categoria and len(lanc.categoria.nome)>23 else (lanc.categoria.nome if lanc.categoria else '-')
        forn = (lanc.fornecedor.nome[:33]+'..') if lanc.fornecedor and len(lanc.fornecedor.nome)>33 else (lanc.fornecedor.nome if lanc.fornecedor else '-')
        
        dados.append([
            lanc.data.strftime('%d/%m/%Y'),
            cat,
            forn,
            lanc.fornecedor.cpf_cnpj if lanc.fornecedor else '-',
            prefixo + 'R$ ' + valor_fmt
        ])
    
    larguras = [2.2*cm, 4.5*cm, 6.5*cm, 3.8*cm, 3*cm]
    tabela = Table(dados, colWidths=larguras, repeatRows=1)
    
    estilo = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), cor_cinza),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 7),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 3),
        ('BOTTOMPADDING', (0,1), (-1,-1), 3),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),
        ('ALIGN', (2,1), (2,-1), 'LEFT'),
        ('ALIGN', (3,1), (3,-1), 'CENTER'),
        ('ALIGN', (4,0), (4,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.Color(0.95,0.95,0.95)])
    ])
    tabela.setStyle(estilo)
    
    for i, lanc in enumerate(lista_final, 1):
        cor = cor_verde if isinstance(lanc, Receita) else cor_vermelha
        tabela.setStyle(TableStyle([('TEXTCOLOR', (4, i), (4, i), cor)]))
    
    # Desenhar com paginação
    tabela.wrapOn(p, width - 2*MARGEM, height)
    altura_total = tabela._height
    
    linhas_por_pag = 24
    linha_atual = 1
    total_linhas = len(dados) - 1
    
    while linha_atual <= total_linhas:
        linhas_nesta = min(linhas_por_pag, total_linhas - linha_atual + 1)
        dados_pag = [dados[0]] + dados[linha_atual:linha_atual + linhas_nesta]
        
        tab_pag = Table(dados_pag, colWidths=larguras)
        tab_pag.setStyle(estilo)
        
        for idx in range(1, len(dados_pag)):
            if linha_atual + idx - 1 <= total_linhas:
                lanc = lista_final[linha_atual + idx - 2]
                cor = cor_verde if isinstance(lanc, Receita) else cor_vermelha
                tab_pag.setStyle(TableStyle([('TEXTCOLOR', (4, idx), (4, idx), cor)]))
        
        tab_pag.wrapOn(p, width - 2*MARGEM, height)
        tab_pag.drawOn(p, MARGEM, y - tab_pag._height)
        
        rodape()
        
        linha_atual += linhas_nesta
        if linha_atual <= total_linhas:
            p.showPage()
            pagina += 1
            y = height - 1 * cm
            y = cabecalho(y, pagina)
            y -= 0.3 * cm
            linhas_por_pag = 24
    
    p.save()
    return response

# ==================== VIEWS DE FORNECEDORES ====================

@login_required
def lista_fornecedores(request):
    """Lista todos os fornecedores do usuário com paginação"""
    # Obter preferências do usuário
    preferencias, _ = PreferenciaUsuario.objects.get_or_create(usuario=request.user)
    itens_por_pagina = int(request.GET.get('per_page', preferencias.itens_por_pagina))
    
    fornecedores = Fornecedor.objects.filter(usuario=request.user).order_by('nome')
    
    # Filtros opcionais
    tipo_filtro = request.GET.get('tipo')
    status_filtro = request.GET.get('status')
    busca = request.GET.get('busca')
    
    if tipo_filtro:
        fornecedores = fornecedores.filter(tipo=tipo_filtro)
    
    if status_filtro == 'ativos':
        fornecedores = fornecedores.filter(ativo=True)
    elif status_filtro == 'inativos':
        fornecedores = fornecedores.filter(ativo=False)
    
    if busca:
        fornecedores = fornecedores.filter(
            Q(nome__icontains=busca) |
            Q(nome_fantasia__icontains=busca) |
            Q(cpf_cnpj__icontains=busca)
        )
    
    # Paginação
    paginator = Paginator(fornecedores, itens_por_pagina)
    page = request.GET.get('page', 1)
    
    try:
        fornecedores_paginados = paginator.page(page)
    except PageNotAnInteger:
        fornecedores_paginados = paginator.page(1)
    except EmptyPage:
        fornecedores_paginados = paginator.page(paginator.num_pages)
    
    context = {
        'fornecedores': fornecedores_paginados,
        'tipo_filtro': tipo_filtro,
        'status_filtro': status_filtro,
        'busca': busca,
        'per_page': itens_por_pagina,
    }
    return render(request, 'APP/fornecedores/lista.html', context)


@login_required
def criar_fornecedor(request):
    """Cria um novo fornecedor"""
    if request.method == 'POST':
        form = FornecedorForm(request.POST)
        if form.is_valid():
            fornecedor = form.save(commit=False)
            fornecedor.usuario = request.user
            fornecedor.save()
            messages.success(request, 'Fornecedor cadastrado com sucesso!')
            return redirect('lista_fornecedores')
    else:
        form = FornecedorForm()
    
    return render(request, 'APP/fornecedores/form.html', {
        'form': form,
        'titulo': 'Novo Fornecedor'
    })


@login_required
def editar_fornecedor(request, pk):
    """Edita um fornecedor existente"""
    fornecedor = get_object_or_404(Fornecedor, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = FornecedorForm(request.POST, instance=fornecedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fornecedor atualizado com sucesso!')
            return redirect('lista_fornecedores')
    else:
        form = FornecedorForm(instance=fornecedor)
    
    return render(request, 'APP/fornecedores/form.html', {
        'form': form,
        'titulo': 'Editar Fornecedor',
        'fornecedor': fornecedor
    })


@login_required
def detalhes_fornecedor(request, pk):
    """Exibe detalhes do fornecedor e suas movimentações"""
    fornecedor = get_object_or_404(Fornecedor, pk=pk, usuario=request.user)
    
    # Buscar despesas e receitas vinculadas
    despesas = fornecedor.despesas.all().order_by('-data')
    receitas = fornecedor.receitas.all().order_by('-data')
    
    # Totais
    total_despesas = despesas.aggregate(Sum('valor'))['valor__sum'] or 0
    total_receitas = receitas.aggregate(Sum('valor'))['valor__sum'] or 0
    
    context = {
        'fornecedor': fornecedor,
        'despesas': despesas,
        'receitas': receitas,
        'total_despesas': total_despesas,
        'total_receitas': total_receitas,
    }
    return render(request, 'APP/fornecedores/detalhes.html', context)


@login_required
def excluir_fornecedor(request, pk):
    """Inativa um fornecedor (soft delete)"""
    fornecedor = get_object_or_404(Fornecedor, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        # Não excluir, apenas inativar
        fornecedor.ativo = False
        fornecedor.save()
        messages.success(request, 'Fornecedor inativado com sucesso!')
        return redirect('lista_fornecedores')
    
    return render(request, 'APP/fornecedores/confirmar_exclusao.html', {
        'fornecedor': fornecedor
    })


# ==================== VIEWS DE EDIÇÃO/EXCLUSÃO DE LANÇAMENTOS ====================

@login_required
def editar_receita(request, pk):
    """Edita uma receita existente"""
    receita = get_object_or_404(Receita, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = ReceitaForm(request.POST, request.FILES, instance=receita, user=request.user)
        if form.is_valid():
            receita = form.save(commit=False)
            
            # Criar fornecedor se preencheu o campo novo_fornecedor
            novo_fornecedor_nome = form.cleaned_data.get('novo_fornecedor')
            if novo_fornecedor_nome:
                fornecedor = Fornecedor.objects.create(
                    usuario=request.user,
                    nome=novo_fornecedor_nome,
                    tipo='PF'
                )
                receita.fornecedor = fornecedor
            
            receita.save()
            messages.success(request, 'Receita atualizada com sucesso!')
            return redirect('listar_lancamentos')
    else:
        form = ReceitaForm(instance=receita, user=request.user)
    
    return render(request, 'APP/receita_form.html', {
        'form': form,
        'receita': receita,
        'editando': True
    })

@login_required
def excluir_receita(request, pk):
    """Exclui uma receita"""
    receita = get_object_or_404(Receita, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        receita.delete()
        messages.success(request, 'Receita excluída com sucesso!')
        return redirect('listar_lancamentos')
    
    return render(request, 'APP/confirmar_exclusao.html', {
        'objeto': receita,
        'tipo': 'receita'
    })

@login_required
def editar_despesa(request, pk):
    """Edita uma despesa existente"""
    despesa = get_object_or_404(Despesa, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = DespesaForm(request.POST, request.FILES, instance=despesa, user=request.user)
        if form.is_valid():
            despesa = form.save(commit=False)
            
            # Criar fornecedor se preencheu o campo novo_fornecedor
            novo_fornecedor_nome = form.cleaned_data.get('novo_fornecedor')
            if novo_fornecedor_nome:
                fornecedor = Fornecedor.objects.create(
                    usuario=request.user,
                    nome=novo_fornecedor_nome,
                    tipo='PJ'
                )
                despesa.fornecedor = fornecedor
            
            despesa.save()
            messages.success(request, 'Despesa atualizada com sucesso!')
            return redirect('listar_lancamentos')
    else:
        form = DespesaForm(instance=despesa, user=request.user)
    
    return render(request, 'APP/despesa_form.html', {
        'form': form,
        'despesa': despesa,
        'editando': True
    })

@login_required
def excluir_despesa(request, pk):
    """Exclui uma despesa"""
    despesa = get_object_or_404(Despesa, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        despesa.delete()
        messages.success(request, 'Despesa excluída com sucesso!')
        return redirect('listar_lancamentos')
    
    return render(request, 'APP/confirmar_exclusao.html', {
        'objeto': despesa,
        'tipo': 'despesa'
    })


# ==================== VIEWS DE DASN-SIMEI ====================

@login_required
def adicionar_dasn_simei(request):
    """Adiciona uma nova declaração DASN-SIMEI"""
    perfil = get_object_or_404(PerfilEmpresa, usuario=request.user)
    
    if request.method == 'POST':
        form = DASN_SIMEIForm(request.POST, request.FILES)
        if form.is_valid():
            dasn = form.save(commit=False)
            dasn.perfil_empresa = perfil
            dasn.save()
            messages.success(request, f'DASN-SIMEI {dasn.ano_calendario} cadastrada com sucesso!')
            return redirect('ver_perfil')
    else:
        form = DASN_SIMEIForm()
    
    return render(request, 'APP/dasn_simei_form.html', {
        'form': form,
        'perfil': perfil
    })


@login_required
def editar_dasn_simei(request, pk):
    """Edita uma declaração DASN-SIMEI existente"""
    perfil = get_object_or_404(PerfilEmpresa, usuario=request.user)
    dasn = get_object_or_404(DASN_SIMEI, pk=pk, perfil_empresa=perfil)
    
    if request.method == 'POST':
        form = DASN_SIMEIForm(request.POST, request.FILES, instance=dasn)
        if form.is_valid():
            form.save()
            messages.success(request, f'DASN-SIMEI {dasn.ano_calendario} atualizada com sucesso!')
            return redirect('ver_perfil')
    else:
        form = DASN_SIMEIForm(instance=dasn)
    
    return render(request, 'APP/dasn_simei_form.html', {
        'form': form,
        'perfil': perfil,
        'dasn': dasn,
        'editando': True
    })


@login_required
def excluir_dasn_simei(request, pk):
    """Exclui uma declaração DASN-SIMEI"""
    perfil = get_object_or_404(PerfilEmpresa, usuario=request.user)
    dasn = get_object_or_404(DASN_SIMEI, pk=pk, perfil_empresa=perfil)
    
    if request.method == 'POST':
        ano = dasn.ano_calendario
        dasn.delete()
        messages.success(request, f'DASN-SIMEI {ano} excluída com sucesso!')
        return redirect('ver_perfil')
    
    return render(request, 'APP/dasn_simei_confirm_delete.html', {
        'dasn': dasn
    })

